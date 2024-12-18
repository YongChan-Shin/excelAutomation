from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import productsData
# import os
# from os import listdir
# from os.path import exists
# from os import makedirs

# 판매중지상품 판매여부 체크용
import excProducts 
excProductsCheck = excProducts.excProducts
excProductsCheckList = []

# 재고정보 생성
wbStock = load_workbook('데이터.xlsx')

stockList = {} # 재고정보
stockErrList = [] # 품절상품 중 판매세팅된 상품정보
stockErrAutoList = [] # 품절상품 중 판매세팅된 상품정보(자동품절건)

soldoutPrdCSList = [] # 품절상품(CS팀전달)

impendingPrdList = [] # 재고 보충 필요 상품정보

first_row_cs = 3
last_row_cs = wbStock['품절상품(CS팀전달)'].max_row + 1

for i in range(first_row_cs, last_row_cs):
  soldoutPrdCSList.append(wbStock['품절상품(CS팀전달)'].cell(i, 17).value)

for wbSheet in wbStock:
  wbFirstCell = 3
  wbLastCell = wbSheet.max_row + 1
  
  for i in range(wbFirstCell, wbLastCell):
    if wbSheet.cell(i, 13).value != None:
      stockList[wbSheet.cell(i, 13).value] = wbSheet.cell(i, 14).value

# 상품정보 리스트
product_list = productsData.product_list
color_list = productsData.color_list
size_list = productsData.size_list

wb = load_workbook('./options.xlsx')
ws = wb.active

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillData2 = PatternFill(fill_type='solid', start_color='FFBDBD', end_color='FFBDBD')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

ws.cell(1, 12).value = '상품정보'
ws.cell(1, 13).value = '상품명'
ws.cell(1, 14).value = '컬러'
ws.cell(1, 15).value = '사이즈'
ws.cell(1, 16).value = '주문정보 정제'
ws.cell(1, 17).value = '재고(데이터파일 기준)'

ws.cell(1, 12).alignment = fillAlignment
ws.cell(1, 13).alignment = fillAlignment
ws.cell(1, 14).alignment = fillAlignment
ws.cell(1, 15).alignment = fillAlignment
ws.cell(1, 16).alignment = fillAlignment
ws.cell(1, 17).alignment = fillAlignment

ws.cell(1, 12).font = fillFont
ws.cell(1, 13).font = fillFont
ws.cell(1, 14).font = fillFont
ws.cell(1, 15).font = fillFont
ws.cell(1, 16).font = fillFont
ws.cell(1, 17).font = fillFont

ws.cell(1, 12).fill = fillData
ws.cell(1, 13).fill = fillData
ws.cell(1, 14).fill = fillData
ws.cell(1, 15).fill = fillData
ws.cell(1, 16).fill = fillData
ws.cell(1, 17).fill = fillData

ws.column_dimensions['L'].width = 40
ws.column_dimensions['M'].width = 25
ws.column_dimensions['N'].width = 25
ws.column_dimensions['O'].width = 25
ws.column_dimensions['P'].width = 40
ws.column_dimensions['Q'].width = 25


first_row = 2
last_row = ws.max_row + 1

for i in range(first_row, last_row):
  try:
    ws.cell(i, 12).value = str(ws.cell(i, 3).value) + '/' + str(ws.cell(i, 4).value) + '/' + str(ws.cell(i, 5).value)
    
    for product in product_list:
      try:
        if product in str(ws.cell(row=i, column=12).value):
          prdDetailInfoProduct = product.replace("(저스틴23)", "")
          ws.cell(row=i, column=13).value = prdDetailInfoProduct
      except:
        pass
    for color in color_list:
      if color in str(ws.cell(row=i, column=12).value):
        prdDetailInfoColor = color
        ws.cell(row=i, column=14).value = prdDetailInfoColor
    for size in size_list:
      if size in str(ws.cell(row=i, column=12).value):
        prdDetailInfoSize = size.replace("FREE", "free")
        ws.cell(row=i, column=15).value = prdDetailInfoSize
    
    prdDetailInfo = '{} {} {}'.format(prdDetailInfoProduct, prdDetailInfoColor, prdDetailInfoSize)
    
    ws.cell(i, 16).value = prdDetailInfo
    ws.cell(i, 17).value = stockList[ws.cell(i, 16).value]
    
    if stockList[ws.cell(i, 16).value] == 0:
      if ws.cell(row=i, column=8).value == "판매중":
        if ws.cell(row=i, column=9).value == "사용":
          if int(ws.cell(row=i, column=7).value) != 0:
            print("{}/{}".format(ws.cell(i, 16).value, stockList[ws.cell(i, 16).value]))
            if ws.cell(i, 16).value in soldoutPrdCSList:
              stockErrList.append("○ {} / 판매상태 : {} / 사용여부 : {} / 재고수량 : {} / 데이터파일 기준 재고 : 0".format(ws.cell(i, 16).value, ws.cell(i, 8).value, ws.cell(i, 9).value, ws.cell(i, 7).value))
            else:
              stockErrAutoList.append("※ 판매량차감 자동품절 상품(CS팀에서 품절로 전달되지 않은 상품) ※\n○ {} / 판매상태 : {} / 사용여부 : {} / 재고수량 : {} / 데이터파일 기준 재고 : 0".format(ws.cell(i, 16).value, ws.cell(i, 8).value, ws.cell(i, 9).value, ws.cell(i, 7).value))
            for colNum in range(1, 18):
              ws.cell(row=i, column=colNum).fill = fillData2
              
    if stockList[ws.cell(i, 16).value] != 0:
      if ws.cell(row=i, column=8).value == "판매중":
        if ws.cell(row=i, column=9).value == "사용":
          if int(ws.cell(row=i, column=7).value) <= 3:
            if stockList[ws.cell(i, 16).value] > int(ws.cell(row=i, column=7).value):
              impendingPrdList.append("○ {} / 판매상태 : {} / 사용여부 : {} / 재고수량 : {} / 데이터파일 기준 재고 : {}".format(ws.cell(i, 16).value, ws.cell(i, 8).value, ws.cell(i, 9).value, ws.cell(i, 7).value, stockList[ws.cell(i, 16).value]))
              
      if ws.cell(row=i, column=8).value == "판매중":
        if ws.cell(row=i, column=9).value == "사용":
          if int(ws.cell(row=i, column=7).value) != 0:              
            if prdDetailInfoProduct in excProductsCheck:
              excProductsCheckList.append("○ {} / 판매상태 : {} / 사용여부 : {} / 재고수량 : {}".format(ws.cell(i, 16).value, ws.cell(i, 8).value, ws.cell(i, 9).value, ws.cell(i, 7).value))
              
  except:
    continue
  
if len(stockErrList) > 0 or len(stockErrAutoList) > 0:
  f = open("(톡스토어) 품절상품 중 판매세팅된 상품 정보.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(톡스토어) 품절상품 중 판매세팅된 상품 정보\n\n")
  for i in stockErrList:
    f.write("{}\n\n".format(i))
  for i in stockErrAutoList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(impendingPrdList) > 0:
  f = open("(톡스토어) 재고 보충 필요 상품 정보(품절 혹은 품절임박).txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(톡스토어) 재고 보충 필요 상품 정보(품절 혹은 품절임박)\n\n")
  for i in impendingPrdList:
    f.write("{}\n\n".format(i))
  f.close()
  
if len(excProductsCheckList) > 0:
  f = open("(톡스토어) 가을 상품 포함 체크.txt", "w")
  f.write("ㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡㅡ\n\n")
  f.write("(톡스토어) 가을 상품 포함 체크\n\n")
  for i in excProductsCheckList:
    f.write("{}\n\n".format(i))
  f.close()  

wb.active.auto_filter.ref = "A1:Q1"
wb.save('상품옵션별 재고현황 추출.xlsx')