from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import productsData
# import os
# from os import listdir
# from os.path import exists
# from os import makedirs

# 재고정보 생성
wbStock = load_workbook('데이터.xlsx')

stockList = {} # 재고정보
stockErrList = [] # 품절상품 중 판매세팅된 상품정보

for wbSheet in wbStock:
  wbFirstCell = 3
  wbLastCell = wbSheet.max_row + 1
  
  for i in range(wbFirstCell, wbLastCell):
    if wbSheet.cell(i, 13).value != None:
      stockList[wbSheet.cell(i, 13).value] = wbSheet.cell(i, 14).value

# 상품정보 리스트
product_list = productsData.product_list
color_list = productsData.color_list
size_list = productsData.size_list

wb = load_workbook('./optionAllList.xlsx')
ws = wb.active

fillData = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
fillData2 = PatternFill(fill_type='solid', start_color='FFBDBD', end_color='FFBDBD')
fillAlignment = Alignment(horizontal='center')
fillFont = Font(bold=True)

ws.cell(1, 18).value = '상품정보'
ws.cell(1, 19).value = '상품명'
ws.cell(1, 20).value = '컬러'
ws.cell(1, 21).value = '사이즈'
ws.cell(1, 22).value = '주문정보 정제'
ws.cell(1, 23).value = '재고(데이터파일 기준)'

ws.cell(1, 18).alignment = fillAlignment
ws.cell(1, 19).alignment = fillAlignment
ws.cell(1, 20).alignment = fillAlignment
ws.cell(1, 21).alignment = fillAlignment
ws.cell(1, 22).alignment = fillAlignment
ws.cell(1, 23).alignment = fillAlignment

ws.cell(1, 18).font = fillFont
ws.cell(1, 19).font = fillFont
ws.cell(1, 20).font = fillFont
ws.cell(1, 21).font = fillFont
ws.cell(1, 22).font = fillFont
ws.cell(1, 23).font = fillFont

ws.cell(1, 18).fill = fillData
ws.cell(1, 19).fill = fillData
ws.cell(1, 20).fill = fillData
ws.cell(1, 21).fill = fillData
ws.cell(1, 22).fill = fillData
ws.cell(1, 23).fill = fillData

ws.column_dimensions['R'].width = 40
ws.column_dimensions['S'].width = 25
ws.column_dimensions['T'].width = 25
ws.column_dimensions['U'].width = 25
ws.column_dimensions['V'].width = 40
ws.column_dimensions['W'].width = 25


first_row = 2
last_row = ws.max_row + 1

# 주문번호 수집
orderDictPrdNums = {}
orderDictPrdDetailNums = {}

for i in range(first_row, last_row):
  try:
    ws.cell(i, 18).value = str(ws.cell(i, 8).value) + '/' + str(ws.cell(i, 11).value)
    
    for product in product_list:
      try:
        if product in str(ws.cell(row=i, column=18).value):
          prdDetailInfoProduct = product.replace("(저스틴23)", "")
          ws.cell(row=i, column=19).value = prdDetailInfoProduct
      except:
        pass
    for color in color_list:
      if color in str(ws.cell(row=i, column=18).value):
        prdDetailInfoColor = color
        ws.cell(row=i, column=20).value = prdDetailInfoColor
    for size in size_list:
      if size in str(ws.cell(row=i, column=18).value):
        prdDetailInfoSize = size.replace("FREE", "free")
        ws.cell(row=i, column=21).value = prdDetailInfoSize
    
    prdDetailInfo = '{} {} {}'.format(prdDetailInfoProduct, prdDetailInfoColor, prdDetailInfoSize)
    
    ws.cell(i, 22).value = prdDetailInfo
    ws.cell(i, 23).value = stockList[ws.cell(i, 22).value]
    
    if stockList[ws.cell(i, 22).value] == 0:
      if ws.cell(row=i, column=16).value == "Y":
        if int(ws.cell(row=i, column=12).value) != 0 or int(ws.cell(row=i, column=13).value) != 0:
          print("{}/{}".format(ws.cell(i, 22).value, stockList[ws.cell(i, 22).value]))
          stockErrList.append("○ {} / 상품코드 : {} / 실재고 : {} / 임시재고 : {} / 사용여부 : {} / 데이터파일 기준 재고 : 0".format(ws.cell(i, 22).value, ws.cell(i, 7).value, ws.cell(i, 12).value, ws.cell(i, 13).value, ws.cell(i, 16).value))
          for colNum in range(1, 24):
            ws.cell(row=i, column=colNum).fill = fillData2
  except:
    continue
  
if len(stockErrList) > 0:
  f = open("보리보리) 품절상품 중 판매세팅된 상품 정보.txt", "w")
  for i in stockErrList:
    f.write("{}\n\n".format(i))
  f.close()

wb.active.auto_filter.ref = "A1:W1"
wb.save('상품옵션별 재고현황 추출.xlsx')